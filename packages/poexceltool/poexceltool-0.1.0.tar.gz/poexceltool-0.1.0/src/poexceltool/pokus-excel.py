import openpyxl
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from tqdm import tqdm

workbook = openpyxl.load_workbook('texts.xlsx')
print(workbook.sheetnames)
barva_1 = Color(theme=0,tint = -0.5)
barva_2 = Color(theme=0,tint = 0)
print(f'složení barvy je: {barva_1}')
workbook['Translations'].freeze_panes = 'B2'
yellow = "00FFFF00"
workbook['Translations'][1][1].fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
print(workbook['Translations'].max_column)
print(workbook['Translations'].max_row)
change = True
for row in tqdm(workbook['Translations'].iter_rows(),position=0,desc='vnější'):
    # print(f"Počet sloupců: {len(row)}",end='\n')
    for cell in tqdm(row, desc='vnitřní',leave=False,position=1):
        if cell.row % 2:
            color = barva_1
        else:
            color = barva_2
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        print(cell.value,end='     ')
    print(end='\n')
for column in workbook['Translations'].iter_cols():
    pomLength = 0
    for cell in column:
        if len(str(cell.value)) > pomLength:
            pomLength = len(str(cell.value))
    workbook['Translations'].column_dimensions[column[0].column_letter].width = min(pomLength,50)
print(workbook['Translations'][1][3].value) #buňka D1
workbook.save('texts.xlsx')