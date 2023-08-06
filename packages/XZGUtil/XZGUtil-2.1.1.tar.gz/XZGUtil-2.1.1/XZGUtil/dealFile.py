#! /usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time    : 2022-06-16 10:07
# @Site    :
# @File    : dealFile.py
# @Software: PyCharm
import os

import openpyxl as op

def save_excel(save_path: str, fileName: str, data: list,sheet_name='Sheet',append=False):
    """
    :param save_path: 路径
    :param fileName: 文件名（不包含后缀）
    :param data: ex:[{"id": 1, "name": "立智", "price": 100}, {"id": 2, "name": "维纳", "price": 200},{"id": 3, "name": "如家", "price": 300}]
    :param sheet_name: sheet名
    :param append: 追加写入
    :return:
    """

    if append:
        fileName = f"{save_path}/{fileName}.xlsx"
        assert os.path.isfile(fileName), f"{fileName} 文件不存在"
        wb = op.load_workbook(fileName)  # 创建工作簿对象
        ws = wb[sheet_name]  # 创建子表
        append_sheet = wb.copy_worksheet(ws)
        rows = append_sheet.max_row
        for row in range(rows + 1, rows + len(data) + 1):
            i = 1
            for l in range(1, len(data[0])+1):
                ws.cell(row, l, list(data[i].values())[l-1])
            i += 1
    else:
        fileName = f"{save_path}/{fileName}.xlsx"
        wb = op.Workbook()  # 创建工作簿对象
        ws = wb['Sheet']  # 创建子表
        title = list(data[0].keys())  # 设置表头
        ws.append(title)  # 添加表头
        for i in range(len(data)):
            d = list(data[i].values())
            ws.append(d)  # 每次写入一行
    wb.save(fileName)


def read_excel(path: str, sheet: str):
    """读取excel"""
    # 打开工作簿
    workbook = op.load_workbook(path)
    # 获取表单
    sheet = workbook[sheet]
    # 获取最大行数
    max_row = sheet.max_row
    # 获取最大列数
    max_column = sheet.max_column
    lr = tuple([tuple([sheet.cell(row=row, column=column).value for column in range(1, max_column+1)]) for row in range(1, max_row+1)])
    print(lr)
    return lr


if __name__ == '__main__':
    testData = [{'表名': 'jd_account_list', '列名': 'id', '数据类型': 'int(64)', '主键': 'YES', '是否可为空': 'NO', '默认值': None, '备注': '', '数据库': 'zein'}, {'表名': 'jd_account_list', '列名': 'platform', '数据类型': 'varchar(64)', '主键': None, '是否可为空': 'YES', '默认值': None, '备注': '平台', '数据库': 'zein'}]
    fileName = '测试4'
    save_excel('./', fileName, testData)
    # read_excel('./测试3.xlsx','Sheet')
